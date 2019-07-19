import xlrd
import openpyxl
import xlwt
from difflib import get_close_matches


class ExcelHandler:
    def __init__(self, filename=None):
        self.workbook = filename

    @property
    def workbook_name(self):
        return self._workbook_name

    @workbook_name.setter
    def workbook_name(self, filename):
        self._workbook_name = filename

    @property
    def workbook(self):
        return self._current_workbook

    @workbook.setter
    def workbook(self, filename = None):
        if filename is None:
            filename = 'database/default.xls'

        self.current_file_name = filename
        _current_workbook = xlrd.open_workbook(self.current_file_name)

        self._current_workbook = openpyxl.Workbook()

        for index in range(_current_workbook.nsheets):
            r_sheet = _current_workbook.sheet_by_index(index)
            w_sheet = self._current_workbook.create_sheet(r_sheet.name, index)
            for row in r_sheet.get_rows():
                updated_row = [col.value for col in row]
                w_sheet.append(updated_row)

        self.sheet = 'Sheet1'

    @property
    def sheet(self):
        return self._current_sheet

    @sheet.setter
    def sheet(self, sheet_name):
        if sheet_name is None:
            sheet_name = 0
        if isinstance(sheet_name, int):
            sheet_name = self.workbook.get_sheet_names[sheet_name]
        if isinstance(sheet_name, str):
            self._current_sheet = self.workbook.get_sheet_by_name(sheet_name)
        else:
            self._current_sheet = None

    def save_workbook(self, filename=None):
        if filename is None:
            filename = self.current_file_name

        if filename.split('.')[-1] == 'xlx':
            workbook = xlwt.Workbook()
            for sheet_name in self.workbook.sheetnames:
                r_sheet = self.workbook.get_sheet_by_name(sheet_name)
                w_sheet = workbook.add_sheet(r_sheet.name)
                for r_index, row in enumerate(r_sheet.get_rows()):
                    for c_index, col in enumerate(row):
                        w_sheet.write(r_index, c_index, col.value)
            workbook.save(filename)
        else:
            self.workbook.save(filename)

    def find_matching_row(self, string):
        rows = []
        for row in self.sheet.rows:
            rows.append(row[0].value)

        # print(rows)
        string = get_close_matches(string, rows, 1, 0.8)[0]
        # print(string)
        return rows.index(string) + 1

    def insert_value(self, row, col, value):
        try:
            return self.sheet.cell(row, col, value)
        except Exception as e:
            print(e)

    def update_excel_file(self, values):
        # print(values)
        row = self.find_matching_row(values['name'])
        # print(row)
        self.insert_value(row, 21, values['index'])
        self.insert_value(row, 22, values['order'])
        self.insert_value(row, 23, values['santoor'])
        print(values['name'], 'Updated Successfully')
